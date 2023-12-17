import csv
import time
import ujson


import undetected_chromedriver as uc
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from datetime import datetime
import openpyxl


def parse_wb():
    driver = uc.Chrome(use_subprocess=False)
    driver.maximize_window()
    links = []
    entrance = int(input("1 - запрос, 2 - ссылка:"))

    if entrance == 1:

        search = input("запрос:")
        pages_count = int(input("Кол-во страниц:"))
        for page in range(1, pages_count + 1):
            driver.get(f'https://www.ozon.ru/search/?text={search}&page={page}&sorting=rating')
            time.sleep(1)
            items = driver.find_elements(By.CLASS_NAME, "product-card__link")
            for item in items:
                links.append(item.get_attribute("href"))
    else:
        search = input("Ссылка:")

        pages_count = int(input("Кол-во страниц:"))
        for page in range(1, pages_count + 1):
            driver.get(f'{search}&page={page}')#https://catalog.wb.ru/catalog/interior3/catalog?TestGroup=no_test&TestID=no_test&appType=1&cat=8459&curr=rub&dest=-1257786&priceU=300000;39841200&sort=rate&spp=27
            time.sleep(0.2)
            try:
                item_data = ujson.loads(driver.find_element(By.TAG_NAME, "body").text)
                item_data = item_data["data"]
                item_data = item_data["products"]
                for i in range(0,len(item_data)):
                    credentials = item_data[i]["supplierId"]
                    links.append(credentials)
            except:
                try:
                    for i in range(1,10):
                        try:
                            driver.get(f'{search}&page={page}')
                            time.sleep(0.2)
                            item_data = ujson.loads(driver.find_element(By.TAG_NAME, "body").text)
                            item_data = item_data["data"]
                            item_data = item_data["products"]
                            for i in range(0, len(item_data)):
                                credentials = item_data[i]["supplierId"]
                                links.append(credentials)
                                break
                        except:
                            pass
                except Exception as a:
                    print(a)

    links = set(links)





    data = [["ОГРН", "ОГРНИП", "ИНН", "Ссылка", "Название"]]

    for link in links:

        driver.get(
            f"https://static-basket-01.wbbasket.ru/vol0/data/supplier-by-id/{link}.json")
        try:
            item_data = ujson.loads(driver.find_element(By.TAG_NAME, "body").text)
        except Exception as a:
            continue

        inn = (item_data)["inn"]
        if inn:
            pass
        else:
            continue

        try:
            ogrn = (item_data)["ogrn"]
        except:
            ogrn = ""
            pass
        try:
            ogrnip = (item_data)["ogrnip"]
        except:
            ogrnip = ""
            pass
        try:
            name_seller = (item_data)["trademark"]
        except:
            name_seller = ""
            pass
        link_seller = (f"https://www.wildberries.ru/seller/{link}")

        data.append([ogrn, ogrnip, inn, link_seller, name_seller])

    with open("data.csv", mode='w', newline='') as file:
        writer = csv.writer(file, delimiter=";")
        writer.writerows(data)
    input()
#https://catalog.wb.ru/catalog/sport13/catalog?TestGroup=no_test&TestID=no_test&appType=1&cat=60141&curr=rub&dest=-1257786&page=1&sort=rate&spp=27








def parse_cdek():
    options = uc.ChromeOptions()
    profile = "C:\\Users\\PC\\AppData\\Local\\Google\\Chrome\\User Data\\Default"
    options.add_argument(f"user-data-dir={profile}")
    driver = uc.Chrome(options=options, use_subprocess=True)
    driver.maximize_window()
    data_itog = [["ИНН", "Ссылка","Можно/Нельзя","Наличие карты","Причина", "Тип Организации", "Название"]]
    time.sleep(2)
    url_cdek = "https://ek5.cdek.ru/page/menuContragent#/viewContragentPage"
    driver.get(url_cdek)
    count_inn = int(input("Введите кол-во строк в документе:"))
    try:
        entrance_cdek = driver.find_element('xpath', '/html/body/div[3]/div/div/form/input[2]')
        entrance_cdek.send_keys(Keys.RETURN)
    except:
        print("мы уже вошли")
    print("мы готовы к работе")
    iframe_first = driver.find_element(By.ID, 'menuContragent_1698203554968')
    driver.switch_to.frame(iframe_first)

    iframe_second = driver.find_element(By.ID, 'gate')
    driver.switch_to.frame(iframe_second)
    data = openpyxl.open("inn.xlsx")
    sheet = data.active
    for i in range(2, count_inn+1):
        inn = int(sheet[f"C{i}"].value)
        link = sheet[f"D{i}"].value
        name_organization = sheet[f"E{i}"].value
        if sheet[f"A{i}"].value:
            type_organization = "ООО"
        else:
            type_organization = "ИП"
        elements = driver.find_elements(By.CLASS_NAME, "floatDigits")
        elements[2].clear()
        elements[2].click()
        elements[2].send_keys(inn)
        elements[2].send_keys(Keys.RETURN)
        time.sleep(1.5)


        try:
            card_absence = driver.find_element(By.XPATH,'/html/body/app-root/div/app-contragent-list-wrapper/div/cdek-resizer/section/div[1]/app-contragent-list-grid/cdek-grid-wrapper/ag-grid-angular/div/div[2]/div[2]/div[6]/div/div')
            not_card = card_absence.find_element(By.CLASS_NAME, 'ag-overlay-loading-center')
            data_itog.append([inn,link,"Можно" ,"1","", type_organization, name_organization])
            print("''''''''''''''''''''''''''''''''")
            print("ИНН:", inn)
            print(not_card.text)
            print("нет карты")
            print("''''''''''''''''''''''''''''''''")
            inn = "0" + str(inn)
            elements = driver.find_elements(By.CLASS_NAME, "floatDigits")
            elements[2].clear()
            elements[2].click()
            elements[2].send_keys(inn)
            elements[2].send_keys(Keys.RETURN)
            time.sleep(1.5)
            card_absence = driver.find_element(By.XPATH,'/html/body/app-root/div/app-contragent-list-wrapper/div/cdek-resizer/section/div[1]/app-contragent-list-grid/cdek-grid-wrapper/ag-grid-angular/div/div[2]/div[2]/div[6]/div/div')
            not_card = card_absence.find_element(By.CLASS_NAME, 'ag-overlay-loading-center')
            data_itog.append([inn,link,"Можно" ,"1","", type_organization, name_organization])
            pass
        except Exception as e:
            pass


        try:
            for i in range(1,50):
                field_data = driver.find_element(By.CLASS_NAME, 'ag-center-cols-container')
                card_data = field_data.find_element(By.XPATH,f'/html/body/app-root/div/app-contragent-list-wrapper/div/cdek-resizer/section/div[1]/app-contragent-list-grid/cdek-grid-wrapper/ag-grid-angular/div/div[2]/div[2]/div[3]/div[2]/div/div/div[{i}]')
                card_data.click()
                time.sleep(1.5)
                contract_data = driver.find_element(By.XPATH,"/html/body/app-root/div/app-contragent-list-wrapper/div/cdek-resizer/section/div[3]/app-contragent-list-details-grid/app-contragent-list-details-contracts/div[2]/cdek-grid-wrapper/ag-grid-angular/div/div[2]/div[2]/div[6]/div/div")
                try:
                    not_date = contract_data.find_element(By.CLASS_NAME, 'ag-overlay-loading-center')
                    print("''''''''''''''''''''''''''''''''")
                    print("ИНН:",inn)
                    print(not_date.text)
                    print("нет договора")
                    print("''''''''''''''''''''''''''''''''")

                    OMP = False
                    OMP = omp_client(driver, OMP)
                    if OMP:
                        data_itog.append([inn, link, "Можно", "0","" ,type_organization, name_organization])

                    else:
                        data_itog.append([inn, link, "Нельзя", "0","Отчёт менеджера продаж"])
                        break

                except:
                    try:
                        date_data = driver.find_elements(By.CLASS_NAME, 'ag-center-cols-container')
                        date_data = date_data[1].find_elements(By.CLASS_NAME, 'ag-cell-value')
                        date_now = datetime.now()
                        proverka = str(date_data[3].text)
                        day_now = int(date_now.strftime("%d"))
                        mounth_now = int(date_now.strftime("%m"))
                        year_now = int(date_now.strftime("%Y"))
                        day, month, year = map(int, proverka.split('.'))
                        if year_now > year and mounth_now >= month or month == 12 and day_now >= day - 10:
                                    print("''''''''''''''''''''''''''''''''")
                                    print("ИНН:", inn)
                                    print(proverka)
                                    print("подходит")
                                    print("''''''''''''''''''''''''''''''''")
                                    dynamics = False
                                    dynamics = dynamics_client(driver, dynamics)
                                    if dynamics:

                                        OMP = False
                                        OMP = omp_client(driver,OMP)
                                        if OMP:
                                            data_itog.append([inn, link, "Можно", "0","", type_organization, name_organization])

                                        else:
                                            data_itog.append([inn, link, "Нельзя", "0","Отчёт менеджера продаж"])
                                            break

                                    else:
                                        data_itog.append([inn, link, "Нельзя", "0", "Дата динамики"])
                                        break



                        else:
                            data_itog.append([inn, link, "Нельзя", "0","Дата договора"])
                            break
                    except Exception:
                        print("''''''''''''''''''''''''''''''''")
                        print("ИНН:", inn)
                        print("нет даты")
                        print("''''''''''''''''''''''''''''''''")
                        dynamics = False
                        dynamics = dynamics_client(driver, dynamics)

                        if dynamics:

                            OMP = False
                            OMP = omp_client(driver, OMP)
                            if OMP:
                                data_itog.append([inn, link, "Можно", "0", "",type_organization, name_organization])

                            else:
                                data_itog.append([inn, link, "Нельзя", "0", "Отчёт менеджера продаж"])
                                break
                        else:
                            data_itog.append([inn, link, "Нельзя", "0","Дата динамики"])
                            break

        except Exception:
            pass
    print(data_itog)
    with open("itog.csv", mode='w', newline='') as file:
        writer = csv.writer(file, delimiter=";")
        writer.writerows(data_itog)

    input()



def dynamics_client(driver,dynamics):
    try:
        dynamics = False
        button = driver.find_element(By.ID, 'openDynamicsModalButton')
        button.click()
        time.sleep(2)
        date_now = datetime.now()
        year_now = int(date_now.strftime("%Y"))
        dynamic_year = driver.find_element(By.XPATH,"/html/body/app-root/cdek-modal/div/div/app-contragent-modal-dynamics/div/div[2]/div[2]/app-dynamics-grid/cdek-grid-wrapper/ag-grid-angular/div/div[2]/div[2]/div[3]/div[2]/div/div/div[2]/div[1]/div/span")
        dynamic_year = int(dynamic_year.text)
        mount_data = driver.find_elements(By.XPATH, '//*[@col-id="month"]' )
        if dynamic_year == year_now-1 and mount_data[2].text != 'Декабрь':
            dynamics = True
            print("год хороший",dynamic_year,year_now)
        if dynamic_year <= year_now-2:
            dynamics = True
            print("год хороший", dynamic_year, year_now)
        close = driver.find_element(By.ID, 'closeByCrossButton')
        close.click()
        return dynamics

    except Exception as e:
        driver.find_element(By.CLASS_NAME,'ek5-icon-cross').click()
        dynamics = True
        return dynamics



def omp_client(driver,OMP):
    try:
        print("мы в проверке на ОМП")
        OMP = False
        button = driver.find_element(By.ID, 'showSMRModalButton')
        button.click()
        time.sleep(1.5)





        date_data = driver.find_elements(By.XPATH, '//*[@col-id="reportDate"]')

        if len(date_data) == 1:
            OMP = True
            close = driver.find_element(By.ID, "closeByCrossButton")
            close.click()
            return OMP

        contact = driver.find_elements(By.XPATH, '//*[@col-id="contactType"]')
        print(contact[1].text)
        if contact[1].text == "Закрывающий отчет":
            OMP = True
            close = driver.find_element(By.ID, "closeByCrossButton")
            close.click()
            return OMP

        date_now = datetime.now()
        day_now = int(date_now.strftime("%d"))
        mounth_now = int(date_now.strftime("%m"))
        year_now = int(date_now.strftime("%Y"))


        local_data = driver.find_elements(By.XPATH, '//*[@col-id="salesManagerName"]')
        manager = local_data[1].text
        manager = manager.split("/")[1]

        for i in range(1,len(date_data)):
            date = date_data[i].text
            date = date[:10]
            day, month, year = map(int, date.split('.'))
            if year <= year_now and month+3 <= mounth_now:
                if year == year_now and month+3 == 12 and day <day_now+10:
                    manager_post = local_data[i].text
                    manager_post = manager_post.split("/")[1]
                    if manager_post == manager:
                        OMP = True
                        print(day, month, year)
                        break
                elif year <= year_now and month+3 < 12:
                    manager_post = local_data[i].text
                    manager_post = manager_post.split("/")[1]
                    if manager_post == manager:
                        OMP = True
                        print(day, month, year)
                        break
                elif year < year_now:
                    manager_post = local_data[i].text
                    manager_post = manager_post.split("/")[1]
                    if manager_post == manager:
                        OMP = True
                        print(day, month, year)
                        break

        close = driver.find_element(By.ID, "closeByCrossButton")
        close.click()
        return OMP
    except Exception as e:
        print(e)
        close = driver.find_element(By.ID, "closeByCrossButton")
        close.click()
        return OMP


def pasians():

    baza = openpyxl.load_workbook("baza.xlsx")
    sheet_baza = baza.active

    itog = openpyxl.load_workbook("itog.xlsx")
    sheet_itog = itog.active
    inns_baza = set(sheet_baza.cell(row=i, column=1).value for i in range(2, sheet_baza.max_row + 1))
    inns_itog = set(sheet_itog.cell(row=j, column=1).value for j in range(2, sheet_itog.max_row + 1))

    # Создание словаря для отслеживания последнего значения каждого ИНН
    inn_last_value = {}

    # Обратный проход по файлу itog.xlsx для удаления дубликатов
    for row in range(sheet_itog.max_row, 1, -1):
        inn_itog = sheet_itog.cell(row=row, column=1).value
        value_itog = sheet_itog.cell(row=row, column=2).value

        # Если ИНН уже есть в словаре, удаляем текущую строку
        if inn_itog in inn_last_value:
            sheet_itog.delete_rows(row)
        else:
            # Обновляем значение в словаре на текущее
            inn_last_value[inn_itog] = value_itog

    inns_to_remove = inns_baza.intersection(inns_itog)


    for row in range(sheet_itog.max_row, 1, -1):
        inn_itog = sheet_itog.cell(row=row, column=1).value
        if inn_itog in inns_to_remove:
            sheet_itog.delete_rows(row)

    # Сохранение изменений
    itog.save("itog_without_duplicates.xlsx")





# def pars_ozon(client: httpx.Client):
#     page = client.get("https://basket-10.wb.ru/vol1381/part138101/138101433/info/sellers.json")
#     soup = bs4.BeautifulSoup(page.text, 'html.parser')
#     print(soup.text)


if __name__ == "__main__":
    print("чё тебе надо, долбаёб")
    print("сдэк вб итог")
    data = input()
    if data == "сдэк":
        parse_cdek()
    if data == "вб":
        parse_wb()
    if data == "итог":
        pasians()
